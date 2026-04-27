# ── Клітинка 2: Запис файлу застосунку ───────────────────────────────
app_code = """
# DEI Index Calculator — MVP Streamlit App
# Digital Ecosystem Intensity for Agri-food Companies
# Based on: Blagodyr et al. (Springer, 2026)

import streamlit as st
import json, os, io
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px

# ── optional deps ─────────────────────────────────────────────────────────────
try:
    import requests, pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    import yfinance as yf
    YF_SUPPORT = True
except ImportError:
    YF_SUPPORT = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# ── constants ─────────────────────────────────────────────────────────────────
DATA_FILE = "dei_data.json"

WEIGHT_SCENARIOS = {
    "Baseline (рівні ваги)":           {"dp": 0.25, "ai": 0.25, "eo": 0.25, "esg": 0.25},
    "Digital-emphasis":                {"dp": 0.35, "ai": 0.30, "eo": 0.20, "esg": 0.15},
    "Ecosystem-emphasis":              {"dp": 0.15, "ai": 0.20, "eo": 0.40, "esg": 0.25},
    "ESG-emphasis":                    {"dp": 0.15, "ai": 0.20, "eo": 0.25, "esg": 0.40},
}

COMPONENTS = {
    "S_DP":  "Digital Platforms",
    "S_AI":  "AI & Precision Agriculture",
    "S_EO":  "Ecosystem Openness",
    "S_ESG": "ESG / Green-Digital",
}

RUBRIC = {
    "S_DP": [
        (0.0, 0.2, "Initial",   "Цифрових платформ немає; ізольовані Excel/1С"),
        (0.2, 0.4, "Basic",     "Окремі цифрові інструменти (ERP або WMS), без системної інтеграції"),
        (0.4, 0.6, "Developed", "Платформа наявна, часткова інтеграція модулів, внутрішнє використання"),
        (0.6, 0.8, "Mature",    "Широке покриття, інтеграція з B2B-партнерами, зовнішній доступ"),
        (0.8, 1.0, "Leader",    "Повна інтеграція, B2B SaaS, зовнішні користувачі платформи"),
    ],
    "S_AI": [
        (0.0, 0.2, "Initial",   "ШІ-ініціатив немає; стандартні цифрові інструменти"),
        (0.2, 0.4, "Basic",     "Окремі пілотні проєкти (NDVI-аналіз, базова аналітика)"),
        (0.4, 0.6, "Developed", "VRA + супутниковий моніторинг + предиктивна аналітика"),
        (0.6, 0.8, "Mature",    "Кілька AI-інструментів по доменах; прогнозування + комп'ютерний зір"),
        (0.8, 1.0, "Leader",    "ML, LLM, широке операційне впровадження (аналог: Kernel DigitalAg)"),
    ],
    "S_EO": [
        (0.0, 0.2, "Initial",   "Повна вертикальна інтеграція, немає зовнішніх зв'язків"),
        (0.2, 0.4, "Basic",     "Окремі партнерські зв'язки, без системної координаційної платформи"),
        (0.4, 0.6, "Developed", "Партнерська програма, цифровий портал для частини партнерів"),
        (0.6, 0.8, "Mature",    "Повноцінний B2B-інтерфейс, моніторинг партнерів, сотні контрагентів"),
        (0.8, 1.0, "Leader",    "Тисячі партнерів, платформа координації (аналог: Kernel Open Agribusiness)"),
    ],
    "S_ESG": [
        (0.0, 0.2, "Initial",   "ESG-показники відсутні, немає зеленої енергетики"),
        (0.2, 0.4, "Basic",     "Окремі сертифікати (ISO), опублікований звіт зі сталості"),
        (0.4, 0.6, "Developed", "Сонячна/біогазова енергетика, екологічні програми"),
        (0.6, 0.8, "Mature",    "Комплексна інтеграція ESG, Carbon Trust або аналогічна сертифікація"),
        (0.8, 1.0, "Leader",    "Кругова економіка + вуглецеве фермерство + висока ESG-рейтингова оцінка"),
    ],
}

KEYWORDS = {
    "S_DP":  ["platform","ERP","SaaS","digital","integrated system","B2B portal",
              "WMS","MES","IoT","цифрова платформа","цифрова трансформація","автоматизація"],
    "S_AI":  ["precision agriculture","AI","machine learning","VRA","satellite",
              "drone","yield prediction","computer vision","точне землеробство",
              "штучний інтелект","супутниковий моніторинг"],
    "S_EO":  ["partner","open agribusiness","ecosystem","contract farming","franchise",
              "partner network","партнер","аутсорсинг","контрактне вирощування"],
    "S_ESG": ["renewable energy","biogas","carbon","ESG","GRI","certification",
              "circular","відновлювана енергія","біогаз","вуглецевий слід","сертифікація"],
}

BENCHMARK = [
    {"company_name":"MHP SE",        "year":2024,"S_DP":0.75,"S_AI":0.55,"S_EO":0.55,"S_ESG":0.60},
    {"company_name":"Kernel Holding","year":2024,"S_DP":0.85,"S_AI":0.85,"S_EO":0.85,"S_ESG":0.85},
    {"company_name":"Astarta-Kyiv",  "year":2024,"S_DP":0.80,"S_AI":0.70,"S_EO":0.65,"S_ESG":0.75},
]

COMPANY_SOURCES = {
    "MHP SE":        {"ir": "https://mhp.com.ua/en/financial-information/annual-reports","ticker":"MHPC.L"},
    "Kernel Holding":{"ir": "https://kernel.ua/investor-relations/financial-reports/",   "ticker":"KER.WA"},
    "Astarta-Kyiv":  {"ir": "https://astartaholding.com/en/investor-relations/",         "ticker":"AST.WA"},
}

LEVEL_COLOR = {
    "Initial":"#e74c3c","Basic":"#e67e22",
    "Developed":"#f39c12","Mature":"#27ae60","Leader":"#1abc9c",
}

# ── core functions ────────────────────────────────────────────────────────────

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE,"r",encoding="utf-8") as f:
            return json.load(f)
    return []

def save_data(data):
    with open(DATA_FILE,"w",encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def calc_dei(scores: dict, weights: dict) -> float:
    return round(
        weights["dp"]  * scores["S_DP"] +
        weights["ai"]  * scores["S_AI"] +
        weights["eo"]  * scores["S_EO"] +
        weights["esg"] * scores["S_ESG"], 4)

def get_level(dei: float) -> str:
    if dei < 0.20: return "Initial"
    if dei < 0.40: return "Basic"
    if dei < 0.60: return "Developed"
    if dei < 0.80: return "Mature"
    return "Leader"

def sensitivity(scores: dict) -> dict:
    return {sc: calc_dei(scores, w) for sc, w in WEIGHT_SCENARIOS.items()}

def score_nlp(text: str) -> dict:
    t = text.lower()
    return {comp: round(min(sum(1 for kw in kws if kw.lower() in t)/len(kws), 1.0), 4)
            for comp, kws in KEYWORDS.items()}

def rubric_tooltip(comp: str) -> str:
    return "\\n".join(f"{lvl} ({lo:.1f}-{hi:.1f}): {desc}"
                     for lo,hi,lvl,desc in RUBRIC[comp])

# ── charts ────────────────────────────────────────────────────────────────────

def radar_chart(records: list) -> go.Figure:
    cats  = list(COMPONENTS.values()) + [list(COMPONENTS.values())[0]]
    keys  = list(COMPONENTS.keys())
    colors= px.colors.qualitative.Set2
    fig   = go.Figure()
    for i, r in enumerate(records):
        vals = [r[k] for k in keys] + [r[keys[0]]]
        fig.add_trace(go.Scatterpolar(
            r=vals, theta=cats, fill="toself",
            name=f"{r['company_name']} ({r['year']})",
            line_color=colors[i % len(colors)], opacity=0.75))
    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0,1])),
        showlegend=True, height=420,
        margin=dict(l=50,r=50,t=40,b=30))
    return fig

def stacked_bar(records: list, scenario: str) -> go.Figure:
    weights = WEIGHT_SCENARIOS[scenario]
    names   = [f"{r['company_name']} ({r['year']})" for r in records]
    keys    = list(COMPONENTS.keys())
    palette = ["#3498db","#2ecc71","#e67e22","#9b59b6"]
    w_list  = list(weights.values())
    fig     = go.Figure()
    for i,(k,label) in enumerate(COMPONENTS.items()):
        fig.add_trace(go.Bar(
            name=f"{label} (×{w_list[i]})",
            x=names, y=[round(r[k]*w_list[i],4) for r in records],
            marker_color=palette[i]))
    fig.update_layout(barmode="stack", height=400,
                      yaxis=dict(range=[0,1],title="DEI"),
                      margin=dict(l=40,r=20,t=40,b=100))
    return fig

# ── exports ───────────────────────────────────────────────────────────────────

def to_csv(data: list) -> str:
    bw = WEIGHT_SCENARIOS["Baseline (рівні ваги)"]
    rows = []
    for r in data:
        sc  = {k:r[k] for k in COMPONENTS}
        dei = calc_dei(sc, bw)
        rows.append({"Компанія":r["company_name"],"Рік":r["year"],
                     **{k:r[k] for k in COMPONENTS},
                     "DEI (Baseline)":dei,"Рівень":get_level(dei)})
    return pd.DataFrame(rows).to_csv(index=False, encoding="utf-8-sig")

def to_excel(data: list) -> bytes:
    if not EXCEL_SUPPORT:
        return b""
    wb = Workbook(); ws = wb.active; ws.title = "DEI Results"
    hdr = ["Компанія","Рік","S_DP","S_AI","S_EO","S_ESG"] + list(WEIGHT_SCENARIOS.keys()) + ["Рівень"]
    ws.append(hdr)
    hfill = PatternFill("solid", fgColor="2C3E50")
    hfont = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill=hfill; cell.font=hfont
        cell.alignment=Alignment(horizontal="center")
    for r in data:
        sc   = {k:r[k] for k in COMPONENTS}
        sens = [calc_dei(sc,w) for w in WEIGHT_SCENARIOS.values()]
        ws.append([r["company_name"],r["year"],
                   r["S_DP"],r["S_AI"],r["S_EO"],r["S_ESG"]] + sens + [get_level(sens[0])])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=18
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ═══════════════════════════════ STREAMLIT UI ═════════════════════════════════

st.set_page_config(
    page_title="DEI Index Calculator",
    page_icon="🌾",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown('''
<style>
.dei-badge{font-size:2.4rem;font-weight:800;padding:8px 22px;
           border-radius:14px;color:#fff;display:inline-block}
.level-tag{font-size:1.15rem;font-weight:600;margin-top:6px}
.sec-hdr{font-size:1rem;font-weight:700;color:#2c3e50;
         border-left:4px solid #3498db;padding-left:8px;margin:14px 0 6px}
</style>
''', unsafe_allow_html=True)

# ── sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🌾 DEI Calculator")
    st.caption("Digital Ecosystem Intensity Index\\nfor Agri-food Companies")
    st.divider()
    page = st.radio("Навігація", [
        "📊 Розрахунок DEI",
        "🗄️ База даних",
        "📈 Порівняння компаній",
        "📥 Збір даних (PDF / API)",
        "ℹ️ Про індекс",
    ])
    st.divider()
    st.caption("Blagodyr et al. Springer, 2026")

data = load_data()

# ══════════════════════════ PAGE 1 — Calculation ══════════════════════════════
if page == "📊 Розрахунок DEI":
    st.header("📊 Розрахунок DEI-індексу")

    col_in, col_out = st.columns([1.1, 1], gap="large")

    with col_in:
        st.markdown('<div class="sec-hdr">1. Дані компанії</div>', unsafe_allow_html=True)
        company_name = st.text_input("Назва компанії", placeholder="напр. MHP SE")
        year = st.selectbox("Рік оцінки", list(range(2025, 2019, -1)))

        st.markdown('<div class="sec-hdr">2. Оцінки компонент  (0.0 – 1.0)</div>',
                    unsafe_allow_html=True)
        scores = {}
        for key, label in COMPONENTS.items():
            scores[key] = st.slider(
                label, 0.0, 1.0, 0.50, 0.05,
                help=rubric_tooltip(key), key=f"s_{key}")

        st.markdown('<div class="sec-hdr">3. Сценарій ваг</div>', unsafe_allow_html=True)
        sc_name = st.selectbox("Сценарій", list(WEIGHT_SCENARIOS.keys()) + ["Custom"])
        if sc_name == "Custom":
            c1,c2,c3,c4 = st.columns(4)
            wd  = c1.number_input("w_DP",  0.0,1.0,0.25,0.05)
            wa  = c2.number_input("w_AI",  0.0,1.0,0.25,0.05)
            we  = c3.number_input("w_EO",  0.0,1.0,0.25,0.05)
            wes = c4.number_input("w_ESG", 0.0,1.0,0.25,0.05)
            weights = {"dp":wd,"ai":wa,"eo":we,"esg":wes}
            wsum = sum(weights.values())
            if abs(wsum-1.0) > 0.001:
                st.warning(f"⚠️ Сума ваг = {wsum:.3f}  (має бути 1.0)")
        else:
            weights = WEIGHT_SCENARIOS[sc_name]

    with col_out:
        dei   = calc_dei(scores, weights)
        level = get_level(dei)
        color = LEVEL_COLOR[level]

        st.markdown('<div class="sec-hdr">Результат</div>', unsafe_allow_html=True)
        st.markdown(
            f'<div class="dei-badge" style="background:{color}">DEI = {dei:.2f}</div>',
            unsafe_allow_html=True)
        st.markdown(
            f'<div class="level-tag" style="color:{color}">Рівень: {level}</div>',
            unsafe_allow_html=True)
        st.divider()

        comp_df = pd.DataFrame({
            "Компонента": list(COMPONENTS.values()),
            "Оцінка":    [scores[k] for k in COMPONENTS],
        })
        st.dataframe(comp_df, use_container_width=True, hide_index=True)

        st.markdown('<div class="sec-hdr">Аналіз чутливості</div>', unsafe_allow_html=True)
        sens = sensitivity(scores)
        sd = pd.DataFrame({"Сценарій":list(sens.keys()),"DEI":list(sens.values())})
        sd["Рівень"] = sd["DEI"].apply(get_level)
        st.dataframe(sd, use_container_width=True, hide_index=True)

    st.divider()
    one_rec = {**scores, "company_name": company_name or "Компанія", "year": year}
    st.plotly_chart(radar_chart([one_rec]), use_container_width=True)

    st.divider()
    ca, cb = st.columns(2)
    with ca:
        if st.button("💾 Зберегти до бази даних", type="primary", use_container_width=True):
            if not company_name.strip():
                st.error("Введіть назву компанії")
            elif (company_name.strip(), year) in {(d["company_name"],d["year"]) for d in data}:
                st.warning("Такий запис вже існує.")
            else:
                data.append({"company_name":company_name.strip(),"year":year,
                             **scores,"scenario":sc_name})
                save_data(data)
                st.success("✅ Збережено!")
    with cb:
        if st.button("📋 Завантажити еталонні дані (MHP, Kernel, Astarta)",
                     use_container_width=True):
            exist = {(d["company_name"],d["year"]) for d in data}
            added = 0
            for b in BENCHMARK:
                if (b["company_name"],b["year"]) not in exist:
                    data.append({**b,"scenario":"Baseline (рівні ваги)"}); added+=1
            save_data(data)
            st.success(f"✅ Додано {added} еталонних записів")

# ══════════════════════════ PAGE 2 — Database ═════════════════════════════════
elif page == "🗄️ База даних":
    st.header("🗄️ База даних компаній")

    if not data:
        st.info("База порожня. Додайте компанії у розділі «Розрахунок DEI».")
    else:
        bw = WEIGHT_SCENARIOS["Baseline (рівні ваги)"]
        rows = []
        for r in data:
            sc  = {k:r[k] for k in COMPONENTS}
            dei = calc_dei(sc, bw)
            rows.append({"Компанія":r["company_name"],"Рік":r["year"],
                         "S_DP":r["S_DP"],"S_AI":r["S_AI"],
                         "S_EO":r["S_EO"],"S_ESG":r["S_ESG"],
                         "DEI":dei,"Рівень":get_level(dei)})
        df = pd.DataFrame(rows).sort_values("DEI",ascending=False)
        st.dataframe(df, use_container_width=True, hide_index=True)

        c1, c2, c3 = st.columns(3)
        with c1:
            st.download_button("⬇️ CSV", to_csv(data), "dei_results.csv",
                               "text/csv", use_container_width=True)
        with c2:
            if EXCEL_SUPPORT:
                st.download_button("⬇️ Excel", to_excel(data), "dei_results.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            else:
                st.caption("pip install openpyxl")
        with c3:
            if st.button("🗑️ Очистити базу", use_container_width=True):
                save_data([]); st.rerun()

# ══════════════════════════ PAGE 3 — Comparison ═══════════════════════════════
elif page == "📈 Порівняння компаній":
    st.header("📈 Порівняння компаній")
    if len(data) < 2:
        st.info("Додайте щонайменше 2 компанії.")
    else:
        labels = [f"{d['company_name']} ({d['year']})" for d in data]
        sel = st.multiselect("Оберіть компанії",labels,default=labels[:min(3,len(labels))])
        sc_bar = st.selectbox("Сценарій для стовпчастої діаграми",
                              list(WEIGHT_SCENARIOS.keys()), key="sc_bar")
        sel_data = [data[labels.index(s)] for s in sel]

        if sel_data:
            c1,c2 = st.columns(2)
            with c1:
                st.subheader("Радарна діаграма")
                st.plotly_chart(radar_chart(sel_data), use_container_width=True)
            with c2:
                st.subheader("Складена стовпчаста діаграма DEI")
                st.plotly_chart(stacked_bar(sel_data, sc_bar), use_container_width=True)

            st.subheader("Таблиця чутливості по сценаріях")
            rows = []
            for r in sel_data:
                sc  = {k:r[k] for k in COMPONENTS}
                row = {"Компанія":f"{r['company_name']} ({r['year']})"}
                row.update({s:calc_dei(sc,w) for s,w in WEIGHT_SCENARIOS.items()})
                rows.append(row)
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

# ══════════════════════════ PAGE 4 — Data collection ══════════════════════════
elif page == "📥 Збір даних (PDF / API)":
    st.header("📥 Автоматичний збір даних із відкритих джерел")

    tab_pdf, tab_api = st.tabs(["📄 PDF-звіт (NLP)", "📡 ESG API (Yahoo Finance)"])

    with tab_pdf:
        st.info("Програма завантажує PDF-звіт, витягує текст і автоматично оцінює компоненти за ключовими словами. Ви можете скоригувати оцінки вручну.")
        known = st.selectbox("Відомі IR-сторінки",
                             ["— вибрати —"]+list(COMPANY_SOURCES.keys()))
        if known != "— вибрати —":
            st.markdown(f"🔗 [Відкрити IR-сторінку]({COMPANY_SOURCES[known]['ir']})  *(скопіюйте URL конкретного PDF нижче)*")
        pdf_url = st.text_input("URL PDF-файлу",
                                placeholder="https://company.com/annual-report-2024.pdf")

        if st.button("🔍 Аналізувати PDF", type="primary"):
            if not PDF_SUPPORT:
                st.error("Встановіть: `pip install requests pdfplumber`")
            elif not pdf_url.strip():
                st.warning("Введіть URL PDF")
            else:
                with st.spinner("Завантажую та аналізую..."):
                    try:
                        os.makedirs("cache", exist_ok=True)
                        fn = os.path.join("cache",
                             pdf_url.split("/")[-1].split("?")[0] or "report.pdf")
                        if not os.path.exists(fn):
                            hdrs = {"User-Agent":"Mozilla/5.0 (compatible; DEI-Bot/1.0)"}
                            r = requests.get(pdf_url, headers=hdrs, timeout=40)
                            r.raise_for_status()
                            with open(fn,"wb") as f: f.write(r.content)
                        with pdfplumber.open(fn) as pdf:
                            text = "\\n".join(p.extract_text() or "" for p in pdf.pages)
                        nlp = score_nlp(text)
                        st.session_state["nlp_scores"] = nlp
                        st.session_state["nlp_chars"]  = len(text)
                        st.success(f"✅ Проаналізовано {len(text):,} символів")
                    except Exception as e:
                        st.error(f"Помилка: {e}")

        if "nlp_scores" in st.session_state:
            nlp = st.session_state["nlp_scores"]
            st.subheader("Авто-оцінки (NLP keyword scoring) — скоригуйте за потреби")
            adj = {}
            for k, label in COMPONENTS.items():
                adj[k] = st.slider(
                    f"{label}  [авто: {nlp[k]:.2f}]",
                    0.0, 1.0, float(nlp[k]), 0.05, key=f"adj_{k}")
            dei_nlp = calc_dei(adj, WEIGHT_SCENARIOS["Baseline (рівні ваги)"])
            st.metric("DEI (Baseline)", f"{dei_nlp:.2f}", get_level(dei_nlp))
            st.divider()
            c_name = st.text_input("Назва компанії", key="nlp_cname")
            c_year = st.selectbox("Рік", list(range(2025,2019,-1)), key="nlp_year")
            if st.button("💾 Зберегти до бази даних", key="save_nlp"):
                if not c_name.strip():
                    st.error("Введіть назву")
                else:
                    data.append({"company_name":c_name.strip(),"year":c_year,
                                 **adj,"scenario":"Baseline (рівні ваги)"})
                    save_data(data); st.success("✅ Збережено!")

    with tab_api:
        st.info("Отримати ESG-оцінку з Yahoo Finance через тикер компанії.")
        ticker_in = st.text_input("Тикер", placeholder="KER.WA або MHPC.L")
        if st.button("📡 Запит ESG"):
            if not YF_SUPPORT:
                st.error("Встановіть: `pip install yfinance`")
            elif not ticker_in.strip():
                st.warning("Введіть тикер")
            else:
                with st.spinner("Запит до Yahoo Finance..."):
                    try:
                        stock = yf.Ticker(ticker_in.strip())
                        esg   = stock.sustainability
                        if esg is None or esg.empty:
                            st.warning("Yahoo Finance не повернув ESG-дані для цього тикера. Перевірте тикер або спробуйте пізніше.")
                        else:
                            st.dataframe(esg, use_container_width=True)
                            if "totalEsg" in esg.index:
                                total = float(esg.loc["totalEsg"].iloc[0])
                                norm  = round(1.0 - min(total/40.0, 1.0), 4)
                                st.metric("S_ESG (нормалізована)", norm,
                                    help="Інвертована шкала Yahoo Finance (нижче = краща ESG)")
                    except Exception as e:
                        st.error(f"Помилка: {e}")

        st.divider()
        st.markdown("#### Відомі IR-сторінки")
        for name, src in COMPANY_SOURCES.items():
            st.markdown(f"- **{name}**: [{src['ir']}]({src['ir']}) — тикер `{src['ticker']}`")

# ══════════════════════════ PAGE 5 — About ════════════════════════════════════
elif page == "ℹ️ Про індекс":
    st.header("ℹ️ Про DEI-індекс")
    st.markdown('''
**Digital Ecosystem Intensity (DEI)** — синтетичний індекс цифрово-екосистемної
інтенсивності агропродовольчих компаній, розроблений у дослідженні:

> Blagodyr L., Karachyna N., Blagodyr A., Babchuk D.
> *«Digital Ecosystems and Value Creation in Ukrainian Agri-food Companies:
> A Multi-Level Comparative Analysis»*, Springer, 2026.

---
### Формула
```
DEI = w_DP × S_DP + w_AI × S_AI + w_EO × S_EO + w_ESG × S_ESG
```
Усі ваги рівні 0.25 у базовому сценарії; підтримуються 4 альтернативних сценарії та Custom.

---
### Компоненти
| Компонента | Зміст |
|---|---|
| S_DP | Цифрові платформи (ERP, SaaS, B2B-портали, IoT) |
| S_AI | ШІ та точне землеробство (VRA, ML, супутниковий моніторинг) |
| S_EO | Відкритість екосистеми (партнерські програми, контрактне фермерство) |
| S_ESG | ESG/Green-Digital (відновлювана енергія, кругова економіка, сертифікації) |

---
### Шкала DEI
| Рівень | Діапазон | Колір |
|---|---|---|
| Initial   | 0.00–0.19 | 🔴 |
| Basic     | 0.20–0.39 | 🟠 |
| Developed | 0.40–0.59 | 🟡 |
| Mature    | 0.60–0.79 | 🟢 |
| Leader    | 0.80–1.00 | 💚 |

---
### Еталонні значення (2024)
| Компанія | S_DP | S_AI | S_EO | S_ESG | DEI | Рівень |
|---|---|---|---|---|---|---|
| MHP SE | 0.75 | 0.55 | 0.55 | 0.60 | **0.61** | Mature |
| Astarta-Kyiv | 0.80 | 0.70 | 0.65 | 0.75 | **0.73** | Mature |
| Kernel Holding | 0.85 | 0.85 | 0.85 | 0.85 | **0.85** | Leader |

---
Розробник застосунку: на основі ТЗ Благодир Л.М., ВНТУ, 2026.
    ''')
"""

with open('dei_app.py', 'w', encoding='utf-8') as f:
    f.write(app_code)
print('✅ dei_app.py збережено:', len(app_code), 'символів');
